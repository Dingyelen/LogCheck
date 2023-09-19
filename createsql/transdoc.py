import pandas as pd
import numpy as np
import openpyxl
import re

from conf.model import ProjectModel


class TransDoc():
    '''
    property:
        common_headers -> list: 公共头
        common_headers_num -> int: 公共头数量
        event_name -> list: 所有事件名
        event_name_col -> int: 事件名所在列
        event_attr_col -> int: 事件属性开始列
        event_attr -> dict
    '''
    def __init__(self, appname: str, transdoc: str) -> str:
        '''
        args:
            appname:项目名称
        '''
        self.appname = appname.upper()
        self.transdoc = transdoc

    def __repr__(self) -> str:
        return f"TransDoc({self.appname})"
    
    def load_transdoc(self):
        try:
            print(self.transdoc)
            workbook = openpyxl.load_workbook(f'./transdoc/{self.transdoc}.xlsx')
            worksheet = workbook.active
        except FileNotFoundError:
            raise FileNotFoundError('''please add the app's excel transdoc in the transdoc path and capitalize the alphabet{self.appname}!!!''')

        return worksheet
    
    @property
    def common_headers(self):
        worksheet = self.load_transdoc()
        common_headers = [cell.value for cell in worksheet[1] if cell.value is not None]

        return common_headers
    
    @property
    def event_name(self):
        workbook = openpyxl.load_workbook(f'./LogCheck/transdoc/{self.appname}.xlsx')
        worksheet = workbook.active
        
        event_col = None
        for cell in worksheet[1]:
            if cell.value == '#event_name':
                event_col = cell.column
                break
        if event_col:
            event_name = []
            for cell in worksheet.iter_rows(min_row=4, values_only=True):
                event_value = cell[event_col-1]
                if event_value is not None:
                    event_name.append(event_value)
        else:
            raise ValueError('''there is not '#event_name' in transdoc''')
        
        return event_name
    
    @property 
    def event_name_col(self):
        return self.common_headers.index('#event_name')+2

    @property 
    def common_headers_num(self):
        return len(self.common_headers)    
    
    @property 
    def event_attr_col(self):
        return self.common_headers_num+2
    
    @property         
    def event_attr(self):
        event_name = self.event_name
        event_name_col = self.event_name_col
        workbook = openpyxl.load_workbook(f'./LogCheck/transdoc/{self.appname}.xlsx')
        worksheet = workbook.active
        event_dict = {}
        for event in event_name:
            row_num = 0
            for row in worksheet.iter_rows(values_only=True):
                cell_value = row[event_name_col-1]
                row_num += 1
                if cell_value == event:
                    event_dict[event] = row_num
        
        return event_dict

    


